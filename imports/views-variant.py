from django.shortcuts import render
from django.views.generic.list import  View
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
import logging
import os.path
from openpyxl import load_workbook
from imports.models import *
from comix.models import Publisher, Series, Issue
from django.utils import timezone
from django.db.models import Q
from decimal import *


logger = logging.getLogger(__name__)


def xscrape_image(gcd_issue_id):
    import urllib.request
    from bs4 import BeautifulSoup
    import re
    from PIL import Image
    from time import sleep

    sleep(1)  # Avoid Error 509 Bandwidth limitation

    # Get cover from "http://www.comics.org/issue/" + gcd_issue_id + '/cover/4/'
    with urllib.request.urlopen('http://www.comics.org/issue/' + str(gcd_issue_id) + '/cover/4/') as response:
        html = response.read()
    soup = BeautifulSoup(html, "html.parser")
    img = soup.find('img', 'cover_img')
    matches = re.match(r'.+src="(http.+/(\d+.jpg)).+', str(img))
    if matches is None:
        return ""  # Missing cover on comics.org
    src_filename = matches.group(1)
    filename = matches.group(2)
    saved_filename = 'comix/static/bigImages/' + filename
    if not os.path.isfile(saved_filename):
        print("Scraping image issue", gcd_issue_id)
        urllib.request.urlretrieve(src_filename, saved_filename)
        sleep(1)
        # Now create thumbnail
        size = (100, 156)
        thumb_filename = 'comix/static/thumbnails/' + filename
        im = Image.open(saved_filename)
        im.thumbnail(size)
        im.save(thumb_filename, "JPEG")
    return filename



def debug(str1, str2=''):
    print(str1, str2)


def make_string(field):
    if field == None:
        return ''
    if isinstance(field, str):
        return field
    else:
        return str(field)


class Comic:
    pass


class XImportExcelView(View):
    template_name = 'imports/imports.html'
    gcd_series_recs = None

    def get(self, request):
        wb = load_workbook(filename="data/PubFix.xlsx")
        wb.guess_types = True
        sheet = wb.active
        cells = sheet['A2':'B62']
        fixes = {}
        messages = [('Catalog No', 'Name', 'Issue', 'Year','Problem')]
        for row in cells:
            fixes[row[0].value] = row[1].value

        wb = load_workbook(filename="data/RTS Master Inv Avail 2016 10 06.xlsx")
        wb.guess_types = True
        sheet = wb.active
        row = 1
        while make_string(sheet['A' + str(row + 1)].value) > '':
            row += 1
            # if row > 6:
            #     break   # ToDo: remove, just for testing
            s_row = str(row)
            Comic.catalog_no = make_string(sheet['A' + s_row].value)
            Comic.publisher = sheet['C' + s_row].value
            Comic.sort_name = sheet['D' + s_row].value
            Comic.name = str(sheet['E' + s_row].value)
            t_year = sheet['F' + s_row].value
            if t_year == None:
                Comic.year = 0
            else:
                Comic.year = int(t_year)
            Comic.vol_no = make_string(sheet['G' + s_row].value)
            Comic.issue = make_string(sheet['H' + s_row].value)
            Comic.issue_text = make_string(sheet['I' + s_row].value)
            Comic.grade = make_string(sheet['K' + s_row].value)
            Comic.price = Decimal(sheet['L' + s_row].value)
            Comic.quantity = int(sheet['M' + s_row].value)
            Comic.si = make_string(sheet['N' + s_row].value)
            Comic.issue_notes = make_string(sheet['J' + s_row].value)
            if Comic.issue_notes > '':
                Comic.issue_notes += ' '
            Comic.issue_notes += make_string(sheet['O' + s_row].value)
            Comic.grade_notes = make_string(sheet['P' + s_row].value)
            Comic.inserts = make_string(sheet['Q' + s_row].value)


            try:
                c_issue = Issue.objects.get(catalog_id = Comic.catalog_no)
                c_issue.price = Comic.price
                c_issue.quantity = Comic.quantity
                c_issue.save()
                # # Temporary fix ToDo: remove
                # c_series = Series.objects.get(id=c_issue.gcd_series_id)
                # if c_series.sort_name != Comic.sort_name:
                #     c_series.sort_name = Comic.sort_name
                #     c_series.save()
                #     print('Set sort_name to ', Comic.sort_name)
                continue  # not necessary, just for clarity

            except ObjectDoesNotExist:
                # add new entry
                debug('creating issue:', Comic.catalog_no)
                issue = Issue()

                # Fill in entries from input. Some may need to be fixed
                # Note: there can be multiple records for the same issue in different grades.
                issue.catalog_id = Comic.catalog_no
                issue.volume = Comic.vol_no
                issue.number = Comic.issue
                issue.issue_text = Comic.issue_text
                issue.notes = Comic.issue_notes
                issue.grade = Comic.grade
                issue.grade_notes = Comic.grade_notes
                issue.image_scanned = 0
                issue.inserts = Comic.inserts
                issue.si = Comic.si
                issue.added_date = timezone.now()
                issue.price = Comic.price
                issue.quantity = Comic.quantity
                issue.status = 'available'

                # Find series from title and year begun
                self.gcd_series_recs = GcdSeries.objects.filter(Q(name=Comic.name) | Q(sort_name=Comic.name))
                self.gcd_series_recs = self.gcd_series_recs.filter(country_id=225, year_began=Comic.year)
                series_ct = len(self.gcd_series_recs)
                debug('series_ct', str(series_ct))
                if series_ct == 0:
                    # failed to get match on name, drop item
                    messages.append((issue.catalog_id, Comic.name, issue.number, Comic.year, "No match on name"))
                    continue
                elif series_ct > 1:
                    # Let's look for a unique match on title and issue #
                    gcd_issue_recs = GcdIssue.objects.filter(Q(series_id__name=Comic.name) |
                                                             Q(series_id__sort_name=Comic.name))
                    gcd_issue_recs = gcd_issue_recs.filter(number=Comic.issue)
                    issue_ct = len(gcd_issue_recs)
                    if issue_ct == 0:
                        messages.append((issue.catalog_id, Comic.name, issue.number, Comic.year,
                                         "No match on issue name and number"))
                        continue
                    elif issue_ct == 1:
                        # assume we found our match
                        gcd_issue = gcd_issue_recs[0]
                        gcd_series = GcdSeries.objects.get(id=gcd_issue.series_id)
                        gcd_publisher = GcdPublisher.objects.get(id=gcd_series.publisher_id)
                    else: # More than one. Try to match based on publisher
                        tmp_issue_recs = gcd_issue_recs.filter(series_id__publisher__name=Comic.publisher)
                        issue_ct = len(tmp_issue_recs)
                        if issue_ct == 0:
                            messages.append((issue.catalog_id, Comic.name, issue.number, Comic.year, "Publisher not found"))
                            continue
                        elif issue_ct > 1: # have variants
                            debug('variants count:', str(issue_ct))
                            gcd_issue = tmp_issue_recs[0]
                            gcd_series = GcdSeries.objects.get(id=gcd_issue.series_id)
                            gcd_publisher = GcdPublisher.objects.get(id=gcd_series.publisher_id)
                            issue.variants = ''
                            for var_issue in tmp_issue_recs:
                                issue.variants += str(var_issue.id) + ',' + str(var_issue.series_id) + ','
                                debug('issue.variants:', issue.variants)
                            messages.append((issue.catalog_id, Comic.name, issue.number, Comic.year,
                                             "Duplicate records for title, issue, publisher"))
                            # We are creating error message but fall through to save issue record.
                            # We won't display this issue because variants are present.
                        else:   # found our match
                            gcd_issue = tmp_issue_recs[0]
                            gcd_series = GcdSeries.objects.get(id=tmp_issue_recs[0].series_id)
                            gcd_publisher = GcdPublisher.objects.get(id=gcd_series.publisher_id)
                else: # just one series_ct
                    gcd_series = self.gcd_series_recs[0]
                    gcd_publisher = GcdPublisher.objects.get(pk=gcd_series.publisher_id)
                    gcd_issue_recs = GcdIssue.objects.filter(number=issue.number, series_id=gcd_series.id)
                    if len(gcd_issue_recs) == 0:
                        a1 = issue.number
                        a2 = gcd_series.id
                        messages.append((issue.catalog_id, Comic.name, issue.number, Comic.year, "Bad issue number?"))
                        continue
                    gcd_issue = gcd_issue_recs[0]
                    # Multiple issues above are variants.
                    if len(gcd_issue_recs) > 1:
                        debug('variants count:', str(len(gcd_issue_recs)))
                        issue.variants = ''
                        for var_issue in gcd_issue_recs:
                            issue.variants += str(var_issue.id) + ',' + str(var_issue.series_id) + ','
                            debug('issue.variants:', issue.variants)
                        messages.append((issue.catalog_id, Comic.name, issue.number, Comic.year,
                                         "Variant issues"))
                        # Again, create error message but keep record

                # Now we have gcd_series

                # Verify Publisher in database or create new
                try:
                    debug('Checking Publisher:', gcd_publisher.name)
                    publisher = Publisher.objects.get(pk=gcd_series.publisher_id)
                except ObjectDoesNotExist:
                    # Need to add publisher.
                    publisher = Publisher(id=gcd_publisher.id, gcd_id=gcd_publisher.id,
                                          name=gcd_publisher.name, issue_ct=gcd_publisher.issue_count)
                    publisher.save()


                # Get or create series record
                try:
                    series = Series.objects.get(pk = gcd_series.id)
                except ObjectDoesNotExist:
                    print("Creating series:", gcd_series.id, gcd_series.name)
                    series = Series()
                    series.id = series.gcd_id = gcd_series.id
                    series.name = gcd_series.name
                    series.sort_name = gcd_series.sort_name
                    series.year_began = gcd_series.year_began
                    series.notes = gcd_series.notes
                    series.issue_count = gcd_series.issue_count
                    series.color = gcd_series.color
                    series.gcd_publisher = publisher
                    series.save()
                    print("Saved series:", series.pk)
                # ToDo: Set genre
                # issue.genre_id =

                issue.gcd_series = series
                issue.publication_date = gcd_issue.publication_date
                issue.gcd_notes = gcd_issue.notes
                issue.gcd_id = gcd_issue.id

                # Now go after cover image
                if issue.cover_image == "":
                    issue.cover_image = scrape_image(issue.gcd_id)

                # success! Save our issue
                issue.save()
                debug("Issue saved", str(issue))
                # tbl_issue.is_done = 'Y'  # Mark record as completed
                # tbl_issue.status = 'available'
                # tbl_issue.save()

        context = {'errors': messages,
                   }
        return render(
            request, self.template_name, context
        )

