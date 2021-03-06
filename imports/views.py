from django.shortcuts import render
from django.views.generic.list import  View
import logging
import os.path
from openpyxl import load_workbook, Workbook
from decimal import *
from .forms import UploadFileForm
from django.http import HttpResponseRedirect
from rts.settings import MEDIA_ROOT
from imports.models import SalesReports
from datetime import datetime
from PIL import Image as PIL_Image
from comix.models import Issue, Image
from django.core.exceptions import ObjectDoesNotExist

logger = logging.getLogger(__name__)
min_custom = 10000000


def scrape_image(gcd_issue_id):
    import urllib.request
    from bs4 import BeautifulSoup
    import re
    from time import sleep

    sleep(2)  # Avoid Error 509 Bandwidth limitation

    # Get cover from "http://www.comics.org/issue/" + gcd_issue_id + '/cover/4/'
    with urllib.request.urlopen('http://www.comics.org/issue/' + str(gcd_issue_id) + '/cover/4/') as response:
        html = response.read()
    soup = BeautifulSoup(html, "html.parser")
    img = soup.find('img', 'cover_img')
    matches = re.match(r'.+src="(http.+/(\d+.jpg)).+', str(img))
    if matches is None:
        return ""  # Missing cover on comics.org
    src_filename = matches.group(1)
    filename = matches.group(2)
    saved_filename = 'comix/static/bigImages/' + filename
    if not os.path.isfile(saved_filename):
        print("Scraping image issue", gcd_issue_id)
        urllib.request.urlretrieve(src_filename, saved_filename)
        sleep(1)
        # Now create thumbnail
        size = (100, 156)
        thumb_filename = 'comix/static/thumbnails/' + filename
        im = PIL_Image.open(saved_filename)
        im.thumbnail(size)
        im.save(thumb_filename, "JPEG")
    return filename



def debug(str1, str2=''):
    print(str1, str2)


def make_string(field):
    if field == None:
        return ''
    if isinstance(field, str):
        return field
    else:
        return str(field)

class ShowSalesView(View):
    template_name = 'imports/sales.html'

    def get(self, request):
        if not request.user.is_staff:
            return HttpResponseRedirect('/login')
        last_report = SalesReports.objects.first()
        return render(request, self.template_name,
                      {'last_report_date': last_report.last_report,}
                      )

    def post(self, request):
        last_report = SalesReports.objects.first()
        startdate = self.request.POST.get('startdate')
        mdy = startdate.split('/')
        try:
            start_date = datetime(int(mdy[2]), int(mdy[0]), int(mdy[1]))
        except:
            return render(request, self.template_name,
                          {'last_report_date': last_report.last_report,
                           'date_error': True
                           }
                          )
        solds = Issue.objects.filter(sold_date__gt=start_date)
        new_sales_report = SalesReports()
        new_sales_report.save()
        return render(request, self.template_name,
                      {'last_report_date': start_date,
                       'success': True,
                       'solds': solds,
                       }
                      )


class ImportExcelView(View):
    template_name = 'imports/imports.html'
    gcd_series_recs = None

    def get(self, request):
        if not request.user.is_staff:
            return HttpResponseRedirect('/login')
        form = UploadFileForm()
        return render(request, self.template_name,
                      {'form': form,
                       'display': True,
                       'type_error': False}
                      )

    def post(self, request):
        form = UploadFileForm(request.POST, request.FILES)
        if not form.is_valid():
            return HttpResponseRedirect('/')

        print(request.FILES['file'].name)
        save_path = os.path.join(MEDIA_ROOT[0], 'imports/' + request.FILES['file'].name)
        print("save_path", save_path)
        with open(save_path, 'wb+') as destination:
            for chunk in request.FILES['file'].chunks():
                destination.write(chunk)

        try:
            wb = load_workbook(filename=save_path)
        except:
            context = {
                       'display': True,
                       'type_error': True,
                       }
            return render(
                request, self.template_name, context
            )

        # wb = load_workbook(filename="data/Updated Inv 2 2017 01 25.xlsx")
        wb.guess_types = True
        sheet = wb.active
        row = 1
        while make_string(sheet['A' + str(row + 1)].value) > '':
            row += 1
            # if row > 6:
            #     break   # ToDo: remove, just for testing
            s_row = str(row)
            if row % 1000 == 0:
                print(s_row)

            catalog_no = make_string(sheet['A' + s_row].value)
            c_issue = Issue.objects.filter(catalog_id=catalog_no)

            if len(c_issue) == 0:
                issue = Issue()
                do_update = False
            else:
                issue = c_issue[0]
                do_update = True

            issue.catalog_id = catalog_no
            issue.title = str(sheet['B' + s_row].value)
            issue.sort_title = make_string(sheet['C' + s_row].value)
            issue.publisher_name = make_string(sheet['D' + s_row].value)
            try:
                issue.year_begun = int(sheet['E' + s_row].value)
            except (TypeError, ValueError):
                issue.year_begun = 0
            issue.volume = make_string(sheet['F' + s_row].value)
            issue.show_number = make_string(sheet['G' + s_row].value)
            try:
                issue.number = int(issue.show_number)
            except (TypeError, ValueError):
                issue.number = 0
            issue.issue_text = make_string(sheet['H' + s_row].value)
            issue.edition = make_string(sheet['I' + s_row].value)
            issue.grade = make_string(sheet['J' + s_row].value)
            issue.price = Decimal(sheet['K' + s_row].value)
            issue.grade_notes = make_string(sheet['O' + s_row].value)
            issue.quantity = int(sheet['L' + s_row].value)
            issue.indicia_date = make_string(sheet['Q' + s_row].value)
            if not do_update:
                issue.image_scanned = 0
            issue.inserts = make_string(sheet['P' + s_row].value)
            issue.scarcity_notes = make_string(sheet['M' + s_row].value)
            issue.notes = make_string(sheet['N' + s_row].value)
            issue.in_gcd_flag = True
            if make_string(sheet['T' + s_row].value) > "":
                issue.in_gcd_flag = False
            issue.numerical_grade = make_string(sheet['U' + s_row].value)
            if issue.numerical_grade == '':
                issue.numerical_grade = 0
            found_series = make_string(sheet['R' + s_row].value)
            try:
                issue.gcd_series_id = int(found_series)
            except (TypeError, ValueError):
                issue.gcd_series_id = 0
            if found_series == '30403':     # Classics Illustrated
                hrn = issue.edition[5:]
                hrn = hrn[:len(hrn)-1]
                try:
                    issue.hrn_number = int(hrn)
                except ValueError:
                    hrn = hrn # no-op
            found_issue = make_string(sheet['S' + s_row].value)
            try:
                issue.gcd_id = int(found_issue)
            except (TypeError, ValueError):
                issue.gcd_id = 0

            # update spreadsheet with other fields
            #   sheet['T' + s_row] = issue.gcd_series.gcd_publisher_id
            #     sheet['U' + s_row] = c_issue.gcd_series.year_began

            # print('.', end='')
            issue.save()
            continue

        # wb.save(filename='data/Updated Inv 3 2017 01 25.xlsx')
        # context = {'errors': messages,
        context = {'errors': [],
                   'record_ct': row - 1,
                   'display': False,
                   'type_error': False,
                              }
        return render(
            request, self.template_name, context
        )

class ProcessImagesView(View):
    template_name = 'process_images.html'
    gcd_series_recs = None

    def get(self, request):
        if not request.user.is_staff:
            return HttpResponseRedirect('/login')
        files_dir = os.path.join(MEDIA_ROOT[0], 'images/')
        image_list = os.listdir(files_dir)
        errors = []
        warnings = []
        processed = []
        for image in image_list:
            if image[0] == '.':
                continue
            source = files_dir + image
            try:
                seq = image.split('_')[1][0]    # will cause "list index out of range" error if not formatted correctly
                print("Saving image for", image)
                issue = Issue.objects.get(catalog_id=image.split('_')[0])
                # remove any existing image reference
                issue.images.filter(file_name__startswith=image.split('_')[0] + image.split('_')[1][0]).delete()
                # also remove any cover images from GCD
                issue.images.filter(is_scanned=False).delete()

                page_filename = 'comix/static/bigImages/' + image
                if os.path.exists(page_filename):
                    os.remove(page_filename)
                os.rename(source, page_filename)
                if seq == '1':
                    # create (or update) thumbnail for cover
                    size = (100, 156)
                    thumb_filename = 'comix/static/thumbnails/' + image
                    im = PIL_Image.open(page_filename)
                    im.thumbnail(size)
                    im.save(thumb_filename, "JPEG")
                new_image = Image(is_scanned=True, file_name=image)
                new_image.save()
                issue.images.add(new_image)
                processed.append(image)
            except IndexError:
                errors.append(image)
            except ObjectDoesNotExist:
                warnings.append(image)

        context = {
            'errors': errors,
            'warnings': warnings,
            'images': processed,
            'process_ct': len(processed),
        }
        return render(
            request, self.template_name, context
        )

